# -*- coding: utf-8 -*-
"""표준국어대사전 '자세히 찾기'에서 구분=관용구 전체를 긁어 엑셀로 떨군다.

엔드포인트: POST https://stdict.korean.go.kr/search/searchDetailWords.do
관용구 필터 = searchWordKindCode=4 (단어0 / 구2 / 관용구4 / 속담5),
어종(searchWordKind)은 전부 켜서 '전체'. 페이지 단위는 pageUnit.

결과 항목 구조(한 항목 = 한 <a class="t_blue1">):
  <a class="t_blue1" onclick="fnSearchView('525010','4')">표제어</a>
  <span class="...hanja_font">한자/원어</span> [관용구]
  <font class="dataLine">뜻풀이</font>
"""
import re
import sys
import time
import pathlib

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

URL = "https://stdict.korean.go.kr/search/searchDetailWords.do"
VIEW = "https://stdict.korean.go.kr/search/searchView.do?word_no={code}&searchKeywordTo=3"
PAGE_UNIT = 50          # 서버가 허용하는 최대(드롭다운 한도)
DELAY = 0.4             # 정부 서버 예의상 페이지 간 대기
MAX_RETRY = 3

# 검색 조건 고정값. searchWordKindCode=4 만 관용구로 켠다.
BASE = [
    ("searchFlag", "Y"), ("searchBoxFlag", "Y"), ("searchType", "1"),
    ("searchOp", "AND"), ("searchTargets", "WORD"), ("searchConditions", "all"),
    ("searchKeyword", ""), ("searchKeywordTo", "3"),
    ("searchWordKindCode", "4"),                       # 관용구
    ("searchWordKind", "1"), ("searchWordKind", "2"),
    ("searchWordKind", "3"), ("searchWordKind", "0"),  # 어종 전체
    ("searchWordKind_all", "-1"),
    ("searchSpCode_all", "-1"), ("searchTechtermCode_all", "-1"),
    ("searchMultimediaCode_all", "-1"), ("searchTargetsOrgLanguage", "-1"),
    ("pageUnit", str(PAGE_UNIT)),
]


def fetch_page(session, page):
    data = BASE + [("pageIndex", str(page))]
    last = None
    for attempt in range(1, MAX_RETRY + 1):
        try:
            r = session.post(URL, data=data, timeout=30)
            r.encoding = "utf-8"
            return r.text
        except requests.RequestException as e:
            last = e
            print(f"  [page {page}] 요청 실패 {attempt}/{MAX_RETRY}: {e}", flush=True)
            time.sleep(1.5 * attempt)
    raise last


def parse_total(soup):
    m = re.search(r"총\s*([\d,]+)\s*개", soup.get_text())
    return int(m.group(1).replace(",", "")) if m else None


def parse_rows(soup):
    rows = []
    for a in soup.find_all("a", class_="t_blue1"):
        oc = a.get("onclick", "")
        m = re.search(r"fnSearchView\('(\d+)'", oc)
        if not m:
            continue
        code = m.group(1)
        headword = a.get_text(" ", strip=True)
        dt = a.find_parent("dt")
        hanja, senses = "", []
        if dt:
            sp = dt.find("span", class_="hanja_font")
            if sp:
                hanja = sp.get_text(" ", strip=True)
            for fo in dt.find_all("font", class_="dataLine"):
                t = fo.get_text(" ", strip=True)
                if t:
                    senses.append(t)
        definition = " / ".join(senses)
        rows.append((code, headword, hanja, definition))
    return rows


def write_xlsx(rows, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "관용구"
    headers = ["번호", "표제어", "원어(한자 등)", "뜻풀이", "사전 링크"]
    ws.append(headers)

    head_fill = PatternFill("solid", fgColor="6B5B4D")
    head_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in ws[1]:
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    for i, (code, head, hanja, definition) in enumerate(rows, start=1):
        link = VIEW.format(code=code)
        ws.append([i, head, hanja, definition, link])
        r = ws.max_row
        ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(r, 2).alignment = Alignment(vertical="top")
        ws.cell(r, 2).font = Font(bold=True)
        ws.cell(r, 3).alignment = Alignment(vertical="top")
        ws.cell(r, 4).alignment = Alignment(wrap_text=True, vertical="top")
        lk = ws.cell(r, 5)
        lk.value = link
        lk.hyperlink = link
        lk.font = Font(color="0563C1", underline="single", size=9)
        for col in range(1, 6):
            ws.cell(r, col).border = border

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 70
    ws.column_dimensions["E"].width = 40
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{ws.max_row}"
    wb.save(out_path)


def main():
    session = requests.Session()
    session.headers["User-Agent"] = "Mozilla/5.0 (PLAY48 idiom harvester)"
    out_path = pathlib.Path(__file__).parent / "korean_idioms.xlsx"

    all_rows, seen = [], set()
    total = None
    page = 1
    while True:
        html = fetch_page(session, page)
        soup = BeautifulSoup(html, "html.parser")
        if total is None:
            total = parse_total(soup)
            pages = (total + PAGE_UNIT - 1) // PAGE_UNIT if total else "?"
            print(f"총 {total}개 / 페이지당 {PAGE_UNIT} -> 약 {pages}페이지", flush=True)
        rows = parse_rows(soup)
        if not rows:
            print(f"[page {page}] 결과 0건 -> 종료", flush=True)
            break
        new = 0
        for row in rows:
            if row[0] in seen:
                continue
            seen.add(row[0])
            all_rows.append(row)
            new += 1
        print(f"[page {page}] {len(rows)}건(신규 {new}) 누적 {len(all_rows)}", flush=True)
        if total and len(all_rows) >= total:
            break
        if len(rows) < PAGE_UNIT:
            break
        page += 1
        time.sleep(DELAY)

    write_xlsx(all_rows, out_path)
    miss = (total - len(all_rows)) if total else 0
    if miss:
        print(f"주의: 총 {total} 중 {len(all_rows)}건 수집(차이 {miss}). README 참고.", flush=True)
    print(f"DONE wrote {len(all_rows)} rows -> {out_path}", flush=True)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:  # 디스패치에서 sentinel 보이게
        print(f"FAILED {type(e).__name__}: {e}", flush=True)
        sys.exit(1)
